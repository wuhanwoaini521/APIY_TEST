# coding=utf-8
import configparser
import os
import sys
import xlrd
import xlwt
import yaml
import json
import openpyxl
from openpyxl.styles import fills

from utils.log_config import MyLogger

current_directory = os.path.dirname(os.path.abspath(__file__))
root_path = os.path.abspath(os.path.dirname(
    current_directory) + os.path.sep + ".")
sys.path.append(root_path)

# 实例化logger
logger = MyLogger().get_logger()


class Handle_Excel:
    """
    Excel读写模块
    """
    def __init__(self, file, sheet_name=None):
        self.file = os.path.join(root_path + os.path.sep + "data" + os.path.sep + file)
        self.sheet_name = sheet_name
        logger.info("☆☆☆读取Excel数据☆☆☆")
        self.wb = openpyxl.load_workbook(self.file)
        self.sheet = self.wb[self.sheet_name]

    def read_sheets(self):
        logger.info("===读取Excel中所有的sheet名称===")
        return self.wb.get_sheet_names()

    def get_row_values(self, row):
        # 获取sheet页中有多少列
        ncols = self.sheet.max_column

        row_data = []
        for i in range(1, ncols-1):
            cell_value = self.sheet.cell(row, i).value
            row_data.append(cell_value)
        return row_data

    def read_excel(self):
        """
        读取excel中用例部分内容
        :return:
        """
        if self.sheet_name:
            sheet = self.wb[self.sheet_name]
            # 获取sheet页中有多少行
            nrows = sheet.max_row
            # 获取sheet页中有多少列
            ncols = sheet.max_column

            print(nrows, ncols)
            if nrows < 6:
                logger.error("该sheet页没有用例存在，请更换用例")
                return False
            excel_list = []
            col_list = []
            # 读取所有内容(用例内容从索引6开始 - 实际表中是第七行)

            config = Config_Operation(root_path + os.path.sep + "configs" + os.path.sep + "config.ini")
            host = config.get_config("test")['host']

            url = host + self.read_url()
            method = self.read_method()
            # excel_list.append((i, url, method, [sheet.iter_cols(1, ncols-2)]))
            # i： 用例的行数， url： 测试路径， method： 测试方法
            for i in range(7, nrows+1):
                excel_list.append((i, url, method, [col_result for col_result in self.get_row_values(i)]))
            return excel_list
        else:
            logger.error("请输入sheet名称！")

    def read_url(self):
        """
        读取用例中url
        :return:
        """
        sheet = self.wb[self.sheet_name]
        return sheet.cell(3, 2).value

    def read_method(self):
        """
        读取用例中的请求方式
        :return:
        """
        sheet = self.wb[self.sheet_name]
        return sheet.cell(4, 2).value


class Write_Excel:
    """
    将txt结果写入到excel的类
    """
    def write_excel(self, file, sheet_name, row, col, result, msg):
        """
        测试结果写入excel
        :param row:
        :param col:
        :param msg:
        :param color:
        :return:
        """
        if "pass" in msg:
            color = "00EC00"
        else:
            color = "FF2D2D"

        file = os.path.join(root_path + os.path.sep + "data" + os.path.sep + file)
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, col).value = msg
        sheet.cell(row, col-1).value = result
        sheet.cell(row, col).fill = fills.GradientFill(stop=(color, color))
        wb.save(file)
        logger.info("数据写入完成！")


class Handle_Yaml:
    """
    读取yaml文件的类
    """
    def __init__(self, file):
        self.file = file
        result = yaml.load(open(file), Loader=yaml.SafeLoader)
        print(json.dumps(result, indent=2))

    def read_yaml(self):
        yaml.load(open(self.file), Loader=yaml.SafeLoader)


class Handle_Json:
    """
    读取json文件的类
    """
    def __init__(self, file):
        self.file = file

    def read_json(self):

        if not os.path.exists(self.file):
            with open(self.file, "w") as f:
                pass

        with open(self.file, "r") as f:
            r = json.load(f)
        return r

    def write_json(self, file, sheet_name, row, col, state):
        # result_json = {f"{file}, f"{sheet_name}": [row, col, state]}
        result_json = {
            file: {sheet_name: {"row": row, "col": col, "state": state}}
        }
        r_json = self.read_json()
        r_json.update(result_json)
        with open(self.file, 'w', encoding="gbk") as f:
            json.dump(r_json, f, indent=4, ensure_ascii=False)

class Handle_Txt:
    """
    测试结果写入到txt文件中，脚本最后会把txt结果写入到excel中
    """
    result_txt = os.path.join(root_path + os.path.sep + "data" + os.path.sep + "result.txt")

    def __init__(self, file=result_txt):
        self.file = file

    def read_txt(self):
        with open(self.file, "r") as f:
            txt_results = f.readlines()
        return txt_results

    def write_txt(self, file, sheet_name, row, col, state, result):
        with open(self.file, "a") as f:
            text = f"{file}@#{sheet_name}@#{row}@#{col}@#{state}@#{result}\n"
            f.write(text)

    def del_txt(self):
        if os.path.exists(self.file):
            os.remove(self.file)



class Config_Operation:
    """
    读取配置文件的类
    """
    def __init__(self, file):
        self.cf = configparser.ConfigParser()
        self.cf.read(file, encoding='utf-8')
        self.file = file

    def get_config(self, section_name):
        """
        获取默认配置
        :return:
        """
        option_dict = {}
        # 读取所有的sections
        sections = self.cf.sections()
        # 获取指定sections下的所有options名
        options = self.cf.options(section_name)
        # print("配置文件中system下的配置：%" % options)
        # 获取指定sections下所有options的键值对
        options_dict = self.cf.items(section_name)
        for option in options_dict:
            k, v = option
            # print("%s: %s" % (k, v))
            option_dict[k] = v
        return option_dict

    def get_section(self):
        """
        获取sections列表
        :return:
        """
        section_list = self.cf.sections()
        return section_list

    def add_section(self, section_name):
        """
        增加配置文件的section_name
        :param section_name:
        :return:
        """
        self.cf.add_section(section_name)
        self.cf.write(open(self.file, mode="w", encoding='utf-8'))

    def add_config(self, section_name, kwargs):
        """
        增加配置文件整体内容
        :param file:
        :return:
        """
        # 根据section_name 增加1
        self.cf.add_section(section_name)
        self.cf[section_name] = kwargs

        self.cf.write(open(self.file, mode="w"))

    def add_config_single(self, section_name, kwargs):
        """
        增加配置文件中某个sections下的options
        :param section_name:
        :param kwargs:
        :return:
        """
        if isinstance(kwargs, dict):
            for k, v in kwargs.items():
                self.cf[section_name][k] = v

        self.cf.write(open(self.file, mode="w", encoding='utf-8'))

    def update_config(self, section_name, kwargs):
        """
        修改配置文件中的内容
        :param section_name:
        :param kwargs:
        :return:
        """
        if isinstance(kwargs, dict):
            for k, v in kwargs.items():
                self.cf.set(section_name, k, v)

        self.cf.write(open(self.file, mode="w", encoding='utf-8'))



if __name__ == '__main__':

    # file_name = "../configs/config.ini"
    # config = Config_Operation(file_name)
    # get_section = config.get_section()
    # print(get_section)
    # get_config = config.get_config("test")
    # print(get_config)

    # ya = Handle_Yaml("../data/env.yml")
    # excel = Handle_Excel("接口测试用例-角色管理.xlsx", "通用接口-2")
    # # excel = Handle_Excel("../data/接口测试用例-角色管理.xls")
    # # result = excel.read_excel()
    # result = excel.read_excel()
    # print(result)

    # json_r = Handle_Txt()
    # # json_r.write_txt("接口测试用例-角色管理", "通用接口-2", 7, 8, "pass")
    # # json_r.read_json()
    # res = json_r.read_txt()
    # print(res)

    # config = Config_Operation("../configs/config.ini")
    # get_section = config.get_config("test")['host']
    # print(get_section)

    print(root_path)